"""Synchronize the static TikTok archive from the public Google Sheet.

The sheet is treated as authoritative for every account that appears in it.
Accounts not present in the sheet are retained from the current data.js so the
archive can be migrated to the master sheet gradually.
"""
from __future__ import annotations

import io
import json
import re
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET_ID = "1C0DP7DKN5QCO5GXdNDWYmvuK8RtGEkYp"
SHEET_XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
ROOT = Path(__file__).resolve().parents[1]
DATA_JS = ROOT / "data.js"

ALLOWED_ACCOUNTS = {
    "istent_theboyz",
    "theboyz_officl",
    "jakeybaee2",
    "kebean.moon",
    "eric.sohn22",
}

HEADER_ALIASES = {
    "date": {"data", "date", "upload date", "data dodania", "date added"},
    "description": {"opis tiktoka", "opis", "description", "caption", "tiktok description"},
    "hashtags": {"hashtagi", "hashtags", "hashtag", "tags"},
    "members": {"czlonkowie", "członkowie", "members", "member", "the boyz members"},
    "tiktok_url": {"link tiktok", "link do tiktoka", "tiktok", "tiktok url", "tiktok link", "url"},
    "drive_url": {
        "link google drive", "google drive", "google drive link", "drive", "drive link",
        "link drive", "link do google drive", "link do pobrania", "download", "download link"
    },
}

MEMBER_CANONICAL = {
    "sangyeon": "Sangyeon", "jacob": "Jacob", "younghoon": "Younghoon",
    "hyunjae": "Hyunjae", "juyeon": "Juyeon", "kevin": "Kevin",
    "new": "New", "q": "Q", "haknyeon": "Haknyeon", "hwall": "Hwall",
    "sunwoo": "Sunwoo", "eric": "Eric", "the boyz": "THE BOYZ", "theboyz": "THE BOYZ",
}


def norm_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    return re.sub(r"\s+", " ", text)


def value_or_hyperlink(cell) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        target = str(cell.hyperlink.target).strip()
        if target.startswith("http"):
            return target
    return str(cell.value or "").strip()


def find_columns(headers: list[str]) -> dict[str, int]:
    found: dict[str, int] = {}
    normalized = [norm_header(h) for h in headers]
    for key, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                found[key] = index
                break
    if "tiktok_url" not in found:
        raise RuntimeError("The sheet must contain a TikTok link column (for example: 'Link TikTok').")
    return found


def parse_date(value: object, tiktok_id: str = "") -> tuple[str, int, str]:
    if isinstance(value, datetime):
        dt = value
    else:
        text = re.sub(r"[^0-9]", "", str(value or ""))
        dt = None
        for fmt in ("%y%m%d", "%Y%m%d"):
            try:
                if len(text) == (6 if fmt == "%y%m%d" else 8):
                    dt = datetime.strptime(text, fmt)
                    break
            except ValueError:
                pass
        if dt is None and tiktok_id.isdigit():
            try:
                timestamp = int(tiktok_id) >> 32
                dt = datetime.utcfromtimestamp(timestamp)
            except Exception:
                dt = None
    if dt is None:
        return "", 0, ""
    return dt.strftime("%Y-%m-%d"), dt.year, dt.strftime("%y%m%d")


def tiktok_id_from_url(url: str) -> str:
    match = re.search(r"/video/(\d+)", url)
    return match.group(1) if match else ""


def account_from_url(url: str) -> str:
    match = re.search(r"tiktok\.com/@([^/?#]+)", url, flags=re.I)
    return match.group(1).strip() if match else ""


def normalize_members(value: object) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    pieces = re.split(r"[,;/|]+", raw)
    output: list[str] = []
    for piece in pieces:
        key = piece.strip().lstrip("@").lower()
        canonical = MEMBER_CANONICAL.get(key, piece.strip())
        if canonical and canonical not in output:
            output.append(canonical)
    return output


def load_previous() -> list[dict]:
    if not DATA_JS.exists():
        return []
    text = DATA_JS.read_text(encoding="utf-8").strip()
    match = re.match(r"^window\.TIKTOK_ARCHIVE_DATA=(.*);\s*$", text, flags=re.S)
    if not match:
        return []
    try:
        payload = json.loads(match.group(1))
    except json.JSONDecodeError:
        return []
    return payload.get("videos", []) if isinstance(payload, dict) else []


def download_sheet() -> bytes:
    request = urllib.request.Request(
        SHEET_XLSX_URL,
        headers={"User-Agent": "tbzarchive-github-sync/2.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        data = response.read()
        content_type = response.headers.get("Content-Type", "")
    if not data.startswith(b"PK"):
        raise RuntimeError(
            "Google did not return an XLSX file. Make sure the spreadsheet is shared as "
            "'Anyone with the link' / viewer. Returned content type: " + content_type
        )
    return data


def parse_sheet(data: bytes, previous_by_id: dict[str, dict]) -> list[dict]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    records: list[dict] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows())
        if not rows:
            continue

        header_row_index = None
        columns = None
        for i, row in enumerate(rows[:20]):
            headers = [str(cell.value or "").strip() for cell in row]
            try:
                candidate = find_columns(headers)
            except RuntimeError:
                continue
            header_row_index = i
            columns = candidate
            break
        if header_row_index is None or columns is None:
            continue

        for row in rows[header_row_index + 1:]:
            def get(key: str) -> str:
                idx = columns.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return value_or_hyperlink(row[idx])

            tiktok_url = get("tiktok_url")
            if "tiktok.com/" not in tiktok_url:
                continue
            tid = tiktok_id_from_url(tiktok_url)
            account = account_from_url(tiktok_url)
            if not tid or account not in ALLOWED_ACCOUNTS:
                continue

            previous = previous_by_id.get(tid, {})
            date, year, date_code = parse_date(get("date"), tid)
            description = get("description") or str(previous.get("description") or "")
            hashtags = get("hashtags") or str(previous.get("hashtags") or "")
            members = normalize_members(get("members")) or previous.get("members") or []
            drive_url = get("drive_url") or str(previous.get("driveUrl") or "")

            records.append({
                "tiktokId": tid,
                "account": account,
                "date": date,
                "dateCode": date_code,
                "year": year,
                "description": description,
                "hashtags": hashtags,
                "members": members,
                "tiktokUrl": tiktok_url,
                "driveUrl": drive_url,
            })

    # Deduplicate, last occurrence wins (useful when a corrected row is appended later).
    unique: dict[str, dict] = {}
    for record in records:
        unique[record["tiktokId"]] = record
    return list(unique.values())


def main() -> None:
    previous = load_previous()
    previous_by_id = {str(row.get("tiktokId") or ""): row for row in previous if row.get("tiktokId")}
    sheet_records = parse_sheet(download_sheet(), previous_by_id)
    if not sheet_records:
        raise SystemExit("No supported TikTok rows were found in the Google Sheet; data.js was not changed.")

    accounts_in_sheet = {row["account"] for row in sheet_records}
    retained = [row for row in previous if row.get("account") not in accounts_in_sheet]
    combined = retained + sheet_records

    # Final de-duplication by TikTok ID.
    unique: dict[str, dict] = {}
    for row in combined:
        tid = str(row.get("tiktokId") or "")
        if tid:
            unique[tid] = row
    videos = list(unique.values())
    videos.sort(key=lambda row: (str(row.get("date") or ""), str(row.get("tiktokId") or "")), reverse=True)

    payload = {
        "sourceSpreadsheet": f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit",
        "videos": videos,
    }
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    serialized = serialized.replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")
    DATA_JS.write_text("window.TIKTOK_ARCHIVE_DATA=" + serialized + ";\n", encoding="utf-8")

    counts: dict[str, int] = {}
    for row in videos:
        counts[row.get("account", "unknown")] = counts.get(row.get("account", "unknown"), 0) + 1
    print(f"Google Sheet rows used: {len(sheet_records)}")
    print(f"Accounts updated from sheet: {', '.join(sorted(accounts_in_sheet))}")
    print(f"Archive total: {len(videos)}")
    for account in sorted(counts):
        print(f"  @{account}: {counts[account]}")


if __name__ == "__main__":
    main()
